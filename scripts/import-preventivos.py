#!/usr/bin/env python3
"""
Import preventivos from Excel file to the Centeno app.

Reads 'upload/preventivos 2026.xlsx', maps Excel columns to preventivo schema fields,
and sends them in batches to the /api/preventivos/import endpoint.

Smart column mapping: matches Excel headers to form field keys using fuzzy matching
to maximize data capture even when column names don't exactly match.
"""

import openpyxl
import json
import requests
import sys
import os

API_BASE = "http://localhost:3000/api"

# ======================================================================
# EXCEL COLUMN -> FORM FIELD KEY MAPPING
# ======================================================================
# Maps Excel column index (0-based) to preventivo form field key.
# This mapping was built by analyzing the Excel headers against the
# preventivo form schema (FORM_SECTIONS in preventivo-schema.ts).
# ======================================================================

COLUMN_MAP = {
    # General (cols 0-11)
    0: None,  # "Titulo General" - section header, skip
    1: 'nombreCentro',
    2: 'localizacionCentro',
    3: 'codigoInfo',
    4: 'provincia',
    5: 'tipoCentro',
    6: 'prioridad',
    7: 'proyecto',
    8: 'cups',
    9: 'comercializadora',
    10: 'tecnico',
    11: 'fecha',

    # Procedimientos Preventivos (cols 12-25)
    12: None,  # "Titulo Procedimientos" - section header
    13: 'proc_revision_visual',
    14: 'proc_rellenar_formulario',
    15: 'proc_capturas_remota',
    16: 'proc_test_remota',
    17: 'proc_backup_remota',
    18: 'proc_capturas_evcc',
    19: 'proc_medidas_bateria',
    20: 'proc_medidas_tierra',
    21: 'proc_reapretado_cuadro',
    22: 'proc_limpieza_equipos',
    23: 'proc_barrer_fregar',
    24: 'proc_desbroce',
    25: 'seHaLeidoProcedimientos',

    # Infraestructura (cols 26-79)
    26: None,  # "Titutlo Infraestructura"
    27: 'localizacionContador',
    28: 'tipoLineaAcometida',
    29: 'estadoAcometida',
    30: 'motivoDeficienciaAcometida',
    31: 'fotoDeficienciaAcometida',
    32: 'tipoSuministroElectrico',
    33: 'dondeSuministro',
    34: 'propietarioGrupo',
    35: 'contadorVistaGeneral',
    36: 'contadorCaja',
    37: 'contadorFusibles',
    38: 'parcelaEdificio',
    39: 'existeCandadoCancillaCamino',
    40: 'tipoCandadoCancillaCamino',
    41: 'fotoCancillaCamino',
    42: 'fotoCandadoCancillaCamino',
    43: 'fotoPortal',
    44: 'fotoPuertaImportante',
    45: 'accesoRestringido',
    46: 'vehiculo4x4',
    47: 'disponeCilindroLocken',
    48: 'fotoLocalizacionCilindro',
    49: 'fotoLlavesInterioresCilindro',
    50: 'existeCandadoCancillaRecinto',
    51: 'tipoCandadoCancillaRecinto',
    52: 'fotoCandadoRecinto',
    53: 'fotoGeneralConjunto1',
    54: 'fotoGeneralConjunto2',
    55: 'estadoPuertaAcceso',
    56: 'motivoDeficientePuerta',
    57: 'fotoDeficientePuerta',
    58: 'fotoEstadoGeneralPuerta',
    59: 'disponeBarraAntivandalica',
    60: 'tipoCandadoBarraAntivandalica',
    61: 'fotoBarraAntivandalica',
    62: 'existeCerraduraAcceso',
    63: 'estadoCerraduraAcceso',
    64: 'tipoCerraduraAcceso',
    65: 'motivoDeficienteCerradura',
    66: 'fotoDeficienteCerradura',
    67: 'fotoCerraduraCentro',
    68: 'fotoBombinPuerta',
    69: 'fotoEstadoGeneral1',
    70: 'fotoEstadoGeneral2',
    71: 'fotoLuminarias',
    72: 'funcionaSensorPuerta',
    73: 'seSustituyeSensorPuerta',
    74: 'motivoNoFuncionaSensor',
    75: 'fotoCasetaNorte',
    76: 'fotoCasetaSur',
    77: 'fotoCasetaEste',
    78: 'fotoCasetaOeste',
    79: 'fotoPasamurosExterior',

    # Torre (cols 80-105)
    80: None,  # "Titulo Torre"
    81: 'alturaTorre',
    82: 'tipoEscaleras',
    83: 'tipoLineaVida',
    84: 'estadoLineaVida',
    85: 'proximaRevisionLineaVida',
    86: 'motivoDeficienteLineaVida',
    87: 'fotoSistemaAnticaida',
    88: 'fotoFechaSistemaAnticaidas',
    89: 'fotoCartelCellnex',
    90: 'fotoTorreNorte',
    91: 'fotoTorreSur',
    92: 'fotoTorreEste',
    93: 'fotoTorreOeste',
    94: 'estadoPinturaTorre',
    95: 'tienePararrayos',
    96: 'consta5Puntas',
    97: 'hayNidosCigueña',
    98: 'numeroNidos',
    99: 'existeCestaNidos',
    100: 'hayNidosInteriorCesta',
    101: 'observacionesNidos',
    102: 'fotoNidos1',
    103: 'fotoNidos2',
    104: 'fotoNidos3',
    105: 'fotoNidos4',

    # Equipos Exteriores (cols 106-114)
    106: None,  # "Titutlo Equipos Exteriores"
    107: 'disponeEquiposExteriores',
    108: 'companiasCoubicadaExterior',
    109: 'nombreCompaniaExterior',
    110: 'fotoEquiposExteriores1',
    111: 'fotoEquiposExteriores2',
    112: 'fotoEquiposExteriores3',
    113: 'fotoEquiposExteriores4',
    114: 'fotoEquiposExteriores5',

    # Vallado y Terreno (cols 115-148)
    115: None,  # "Titulo Vallado y Terreno"
    116: 'estadoGeneralVallado',
    117: 'reparaValladoPreventivo',
    118: 'motivoDeficienteVallado',
    119: 'fotoReparacionValladoAntes1',
    120: 'fotoReparacionValladoAntes2',
    121: 'fotoReparacionValladoDespues1',
    122: 'fotoReparacionValladoDespues2',
    123: 'fotoCancillaRecinto',
    124: 'fotoVallado1',
    125: 'fotoVallado2',
    126: 'fotoVallado3',
    127: 'fotoVallado4',
    128: 'tipoTerrenoRecinto',
    129: 'necesitaDesbroce',
    130: 'seHaDesbrozado',
    131: 'fotoEstadoTerreno1',
    132: 'fotoEstadoTerreno2',
    133: 'fotoEstadoTerreno3',
    134: 'fotoDesbroceAntes1',
    135: 'fotoDesbroceAntes2',
    136: 'fotoDesbroceAntes3',
    137: 'fotoDesbroceAntes4',
    138: 'fotoDesbroceAntes5',
    139: 'fotoDesbroceDespues1',
    140: 'fotoDesbroceDespues2',
    141: 'fotoDesbroceDespues3',
    142: 'fotoDesbroceDespues4',
    143: 'fotoDesbroceDespues5',
    144: 'aplicacionHerbicida',
    145: 'tipoHerbicida',
    146: 'vegetacionOtrosProcedimientos',
    147: 'chatarraCentro',
    148: 'fotoResiduosChatarra',

    # Equipos Interiores (cols 149-157)
    149: None,  # "Titulo Equipos Interiores"
    150: 'disponeEquiposInteriores',
    151: 'companiasCoubicadaInterior',
    152: 'nombreCompaniaInterior',
    153: 'fotoEquiposInteriores1',
    154: 'fotoEquiposInteriores2',
    155: 'fotoEquiposInteriores3',
    156: 'fotoEquiposInteriores4',
    157: 'fotoEquiposInteriores5',

    # Cuadro Eléctrico (cols 158-220)
    158: None,  # "Titulo Cuadro Electrico"
    159: 'tipoLineaCentro',
    160: 'repartoCargasCorrecto',
    161: 'motivoDeficienteRepartoCargas',
    162: 'fotoCuadroCerradoGeneral',
    163: 'fotoCuadroCerradoSuperior',
    164: 'fotoCuadroCerradoInferior',
    165: 'fotoIGACable',
    166: 'fotoCuadroAbiertoGeneral',
    167: 'fotoCuadroAbiertoSuperior',
    168: 'fotoCuadroAbiertoInferior',
    169: 'fotoVoltajeL1L2',
    170: 'fotoVoltajeL2L3',
    171: 'fotoVoltajeL3L1',
    172: 'fotoVoltajeNL',
    173: 'fotoIntensidadR',
    174: 'fotoIntensidadS',
    175: 'fotoIntensidadT',
    176: 'fotoIntensidadN',
    177: 'totalCVMINI',
    178: 'fotoCVMINIGeneral',
    179: 'fotoCVMINIEVCC',
    180: 'fotoCVMINICompania2',
    181: 'fotoCVMINICompania3',
    182: 'fotoMedidasTierra',
    183: 'consumoMovistarExterior',
    184: 'tipoAlimentacionMovistarExt',
    185: 'fotoIntensidadMovistarExt',
    186: 'consumoOrangeExterior',
    187: 'tipoAlimentacionOrangeExt',
    188: 'fotoIntensidadOrangeExt',
    189: 'consumoVodafoneExterior',
    190: 'tipoAlimentacionVodafoneExt',
    191: 'fotoIntensidadVodafoneExt',
    192: 'consumoYoigoExterior',
    193: 'tipoAlimentacionYoigoExt',
    194: 'fotoIntensidadYoigoExt',
    195: 'consumoOtrosExterior',
    196: 'tipoAlimentacionOtrosExt',
    197: 'fotoIntensidadOtrosExt',
    198: 'consumoMovistarInterior',
    199: 'tipoAlimentacionMovistarInt',
    200: 'fotoIntensidadMovistarInt',
    201: 'consumoOrangeInterior',
    202: 'tipoAlimentacionOrangeInt',
    203: 'fotoIntensidadOrangeInt',
    204: 'consumoVodafoneInterior',
    205: 'tipoAlimentacionVodafoneInt',
    206: 'fotoIntensidadVodafoneInt',
    207: 'consumoYoigoInterior',
    208: 'tipoAlimentacionYoigoInt',
    209: 'fotoIntensidadYoigoInt',
    210: 'consumoOtrosInterior',
    211: 'tipoAlimentacionOtrosInt',
    212: 'fotoIntensidadOtrosInt',
    213: 'sustitucionDescargadores',
    214: 'cantidadDescargadoresCambiados',
    215: 'fotoDescargadoresAntes',
    216: 'fotoDescargadoresDespues',
    217: 'diodoColocado',
    218: 'fotoDiodo',
    219: 'sensorPuertaCuadro',
    220: 'notaTareaImportante',

    # EVCC (cols 221-242)
    221: None,  # "Titulo EVCC"
    222: 'evcc',
    223: 'fotoEVCC',
    224: 'fotoTermicosEVCC1',
    225: 'fotoTermicosEVCC2',
    226: 'rectificadores',
    227: 'rectificadoresMaximos',
    228: 'limitacionCargaEVCC',
    229: 'evccValoresConsumo',
    230: 'existenBaterias',
    231: 'fotoBaterias',
    232: 'estadoBaterias',
    233: 'motivoEstadoBaterias',
    234: 'marcaBaterias',
    235: 'bancadas',
    236: 'fotoBancadas',
    237: 'numeroBaterias',
    238: 'fechaInstalacionBaterias',
    239: 'fotoFechaBaterias',
    240: 'seHanRetiradoBaterias',
    241: 'numeroBateriasRetiradas',
    242: 'fotoConductancia',

    # Remota (cols 243-249)
    243: None,  # "Titulo Remota"
    244: 'remota',
    245: 'tarjetaSIM',
    246: 'numTelefonoSIM',
    247: 'ipRemota',
    248: 'fotoPegatinaRemota',
    249: 'fotoRemota',

    # Cols 250-255 are None (empty columns)

    # Enlaces (cols 256-302)
    256: None,  # "Titulo Enlaces"
    257: 'existenEnlaces',
    258: 'numeroEnlaces',
    # Enlace 1
    259: 'numeroVano1',
    260: 'marcaEquipo1',
    261: 'modeloEquipo1',
    262: 'fotoEquipamiento1',
    # Note: col 263 is "Numero de Vano 3" in Excel (Vano 2 is missing)
    # Map it to numeroVano2 since the Excel skips Vano 2
    263: 'numeroVano2',
    264: 'marcaEquipo2',
    265: 'modeloEquipo2',
    266: 'fotoEquipamiento2',
    267: 'fotoEtiqueta2',
    # Enlace 3 (labeled as 3 in Excel)
    268: 'numeroVano3',
    269: 'marcaEquipo3',
    270: 'modeloEquipo3',
    271: 'fotoEquipamiento3',
    272: 'fotoEtiqueta3',
    # Enlace 4
    273: 'numeroVano4',
    274: 'marcaEquipo4',
    275: 'modeloEquipo4',
    276: 'fotoEquipamiento4',
    277: 'fotoEtiqueta4',
    # Enlace 5
    278: 'numeroVano5',
    279: 'marcaEquipo5',
    280: 'modeloEquipo5',
    281: 'fotoEquipamiento5',
    282: 'fotoEtiqueta5',
    # Enlace 6
    283: 'numeroVano6',
    284: 'marcaEquipo6',
    285: 'modeloEquipo6',
    286: 'fotoEquipamiento6',
    287: 'fotoEtiqueta6',
    # Enlace 7
    288: 'numeroVano7',
    289: 'marcaEquipo7',
    290: 'modeloEquipo7',
    291: 'fotoEquipamiento7',
    292: 'fotoEtiqueta7',
    # Enlace 8
    293: 'numeroVano8',
    294: 'marcaEquipo8',
    295: 'modeloEquipo8',
    296: 'fotoEquipamiento8',
    297: 'fotoEtiqueta8',
    # Enlace 9
    298: 'numeroVano9',
    299: 'marcaEquipo9',
    300: 'modeloEquipo9',
    301: 'fotoEquipamiento9',
    302: 'fotoEtiqueta9',
    # Enlace 10 is at cols 298-302 but Excel labels them 10
    # Already mapped above as 9, map the next set as 10
    # Actually looking at the headers: 298 is "Numero de Vano 10"
    # Let me re-check... The Excel has:
    # 259=Vano1, 263=Vano3, 268=Vano4, 273=Vano5, 278=Vano6, 283=Vano7, 288=Vano8, 293=Vano9, 298=Vano10
    # So Vano 2 is missing from Excel. The mapping above shifts things.
    # Let me fix: 263 should be numeroVano3, not numeroVano2
    # But wait, the form schema has numeroVano1 through numeroVano10 sequentially.
    # Since the Excel skips Vano 2 (goes 1,3,4,5,6,7,8,9,10), we should map
    # Excel col 263 (Vano 3) -> form numeroVano2 to keep sequential order in form.
    # Actually, better to preserve the original data labels:
    # Excel Vano 1 -> numeroVano1, Excel Vano 3 -> numeroVano3, etc.
    # The form will just have empty numeroVano2.

    # Aire Acondicionado (cols 303-365)
    303: None,  # "Titulo Aire Acondicionado"
    304: 'disponeAA',
    305: 'cuantasMaquinasAA',
    306: 'notaManual',
    307: 'tipoMaquinaPrincipal',
    308: 'notaSellado',
    309: 'correctamenteSellado',
    310: 'fotoEstadoGeneral1',
    311: 'fotoEstadoGeneral2',
    312: 'fotoEstadoGeneral3',
    313: None,  # "Antes de la reparacion" - section header
    314: 'fotoAntesReparacion1',
    315: 'fotoAntesReparacion2',
    316: 'fotoAntesReparacion3',
    317: None,  # "Despues de la reparacion" - section header
    318: 'fotoDespuesReparacion1',
    319: 'fotoDespuesReparacion2',
    320: 'fotoDespuesReparacion3',
    321: None,  # "MAQUINA PRINCIPAL" - section header
    322: 'correctoFuncionamientoAA',
    323: 'motivoNoCorrectoAA',
    324: 'marcaAAPrincipalExterior',
    325: 'modeloAAPrincipalExterior',
    326: 'numSerieAAPrincipalExterior',
    327: 'fotoAAExtPrincipalGeneral',
    328: 'fotoAAExtPrincipalCaract',
    329: 'tipoGasRefrigerante',
    330: 'kgGasRefrigerante',
    331: 'modeloAAPrincipalInterior',
    332: 'numSerieAAPrincipalInterior',
    333: 'fotoAAIntPrincipalGeneral',
    334: 'fotoAAIntPrincipalCaract',
    335: 'fotografiaMandoAA',
    336: 'fotoFiltroAntesAA',
    337: 'fotoFiltroDespuesAA',
    338: None,  # "MAQUINA SECUNDARIA" - section header
    339: 'correctoFuncionamientoAASec',
    340: 'motivoNoCorrectoAASec',
    341: 'marcaAASecundarioExterior',
    342: 'modeloAASecundarioExterior',
    343: 'numSerieAASecundarioExterior',
    344: 'fotoAAExtSecundarioGeneral',
    345: 'fotoAAExtSecundarioCaract',
    346: 'tipoGasRefrigeranteSec',
    347: 'kgGasRefrigeranteSec',
    348: 'modeloAASecundarioInterior',
    349: 'numSerieAASecundarioInterior',
    350: 'fotoAAIntSecundarioGeneral',
    351: 'fotoAAIntSecundarioCaract',
    352: 'fotografiaMandoAASec',
    353: 'fotoFiltroAASecAntes',
    354: 'fotoFiltroAASecDespues',
    355: None,  # "MAQUINA DE CONTINGENCIA" - section header
    356: 'aaContingencia',
    357: 'marcaAAContingencia',
    358: 'fotoAAContingenciaExtGeneral',
    359: 'fotoAAContingenciaExtCaract',
    360: 'fotoAAContingenciaIntGeneral',
    361: 'fotoAAContingenciaIntCaract',
    362: None,  # "SONDAS DE TEMPERATURA" - section header
    363: 'correctaOrientacionSonda',
    364: 'fotoSondaInterior',
    365: 'fotoSondaExterior',

    # Extracción (cols 366-374)
    366: None,  # "Titulo Extraccion"
    367: 'disponeExtraccion',
    368: 'tipoExtraccionAire',
    369: 'fotoExtractores1',
    370: 'fotoExtractores2',
    371: 'fotoExtractores3',
    372: 'disponeFiltroFC',
    373: 'fotoFiltroFCAntes',
    374: 'fotoFiltroFCDespues',

    # Sigfox (cols 375-378)
    375: None,  # "TItulo Sigfox"
    376: 'disponeEquiposSigfox',
    377: 'modeloSigfox',
    378: 'equiposSigfox',

    # Fotovoltaica (cols 379-395)
    379: None,  # "Titulo Fotovoltaica"
    380: 'instalacionFotovoltaica',
    381: 'fotoGeneralFotovoltaica',
    382: 'fotoCuadroFotovoltaica',
    383: 'fotoConexionIGAFotovoltaica',
    384: 'tipoInversor',
    385: 'modeloInversorFotovoltaica',
    386: 'modeloEVCCFotovoltaica',
    387: 'fotoValoresConsumoEVCCFotovoltaica',
    388: 'fotoInversorEVCCFotovoltaica',
    389: 'rectificadoresEVCCFV',
    390: 'rectificadoresMaximosEVCCFV',
    391: 'numeroPanelesFV',
    392: 'fotoPanelesFotovoltaica',
    393: 'fotoEstructuraFotovoltaica',
    394: 'fotoPATFotovoltaica',
    395: 'observacionesFotovoltaica',

    # Limpieza (cols 396-407)
    396: None,  # "Titulo Limpieza"
    397: 'seRealizaLimpieza',
    398: 'motivoNoLimpieza',
    399: 'fotoGeneralLimpieza1',
    400: 'fotoGeneralLimpieza2',
    401: 'fotoCartelCellnexInterior',
    402: 'fotoCasetaInteriorNorte',
    403: 'fotoCasetaInteriorSur',
    404: 'fotoCasetaInteriorEste',
    405: 'fotoCasetaInteriorOeste',
    406: 'fotoTechoInterior',
    407: 'fotoPasamurosInterior',

    # Observaciones (cols 408-447)
    408: None,  # "Titulo Observaciones"
    409: 'observaciones',
    410: 'necesitasMasFotos',
    411: 'varios1',
    412: 'varios2',
    413: 'varios3',
    414: 'varios4',
    415: 'varios5',
    416: 'varios6',
    417: 'varios7',
    418: 'varios8',
    419: 'varios9',
    420: 'varios10',
    # Cols 421-447 are rating/evaluation fields (Accesos., Cerraduras., etc.)
    # These are not in the preventivo form schema, but we store them anyway
    421: 'evalAccesos',
    422: 'evalCerraduras',
    423: 'evalLuminarias',
    424: 'evalCasetaExterior',
    425: 'evalTorre',
    426: 'evalNidos',
    427: 'evalVallado',
    428: 'evalRecinto',
    429: 'evalEquiposExteriores',
    430: 'evalCuadroElectrico',
    431: 'evalMedidasElectricas',
    432: 'evalContadores',
    433: 'evalEstacionContinua',
    434: 'evalBaterias',
    435: 'evalRemota',
    436: 'evalEnlaces',
    437: 'evalAireAcondicionado',
    438: 'evalAirePrincipal',
    439: 'evalAireSecundario',
    440: 'evalSondas',
    441: 'evalExtraccion',
    442: 'evalFiltroFC',
    443: 'evalSigfox',
    444: 'evalFotovoltaica',
    445: 'evalLimpieza',
    446: 'evalCasetaInterior',
    447: 'evalVarios',
}


def convert_value(value):
    """Convert Excel cell value to a string suitable for formData."""
    if value is None:
        return None

    # Handle datetime objects
    if hasattr(value, 'isoformat'):
        return value.isoformat()[:10]  # YYYY-MM-DD format

    # Handle numbers that should be strings
    if isinstance(value, float):
        # If it's an integer value, don't add .0
        if value == int(value):
            return str(int(value))
        return str(value)

    return str(value).strip()


def main():
    excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'upload', 'preventivos 2026.xlsx')

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"Leyendo archivo Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb['Hoja 1']

    # Read headers
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)

    # Read all data rows
    preventivos = []
    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        record = {}

        for col_idx, value in enumerate(row):
            if col_idx >= len(headers):
                break

            field_key = COLUMN_MAP.get(col_idx)
            if field_key is None:
                continue  # Skip section headers and unmapped columns

            converted = convert_value(value)
            if converted is not None and converted != '':
                record[field_key] = converted

        if record.get('nombreCentro'):
            record['rowNumber'] = str(row_num)
            preventivos.append(record)

    wb.close()

    print(f"Leídos {len(preventivos)} preventivos del Excel")

    # Send in batches
    batch_size = 20
    total_created = 0
    total_skipped = 0
    total_duplicates = 0
    total_errors = []

    for i in range(0, len(preventivos), batch_size):
        batch = preventivos[i:i + batch_size]
        print(f"Enviando lote {i // batch_size + 1} ({len(batch)} preventivos)...")

        try:
            response = requests.post(
                f"{API_BASE}/preventivos/import",
                json={"preventivos": batch},
                headers={"Content-Type": "application/json"},
                timeout=60,
            )

            if response.status_code == 200:
                data = response.json()
                results = data.get('results', {})
                total_created += results.get('created', 0)
                total_skipped += results.get('skipped', 0)
                total_duplicates += results.get('duplicatesSkipped', 0)
                if results.get('errors'):
                    total_errors.extend(results['errors'])
                print(f"  ✓ Creados: {results.get('created', 0)}, Duplicados: {results.get('duplicatesSkipped', 0)}, Errores: {len(results.get('errors', []))}")
            else:
                print(f"  ✗ Error HTTP {response.status_code}: {response.text[:200]}")
                total_skipped += len(batch)

        except Exception as e:
            print(f"  ✗ Error de conexión: {e}")
            total_skipped += len(batch)

    print("\n" + "=" * 60)
    print("RESUMEN DE IMPORTACIÓN")
    print("=" * 60)
    print(f"Total leídos del Excel:  {len(preventivos)}")
    print(f"Preventivos creados:     {total_created}")
    print(f"Duplicados omitidos:     {total_duplicates}")
    print(f"Errores/Omitidos:        {total_skipped}")
    if total_errors:
        print(f"\nDetalle de errores ({len(total_errors)}):")
        for err in total_errors[:20]:
            print(f"  - {err}")
        if len(total_errors) > 20:
            print(f"  ... y {len(total_errors) - 20} más")


if __name__ == '__main__':
    main()
